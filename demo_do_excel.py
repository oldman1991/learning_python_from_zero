from openpyxl import load_workbook


def read_excel(path):
    """
    读取excel
    :param path: excel的路径
    :return:
    """
    wb = load_workbook(path)
    # 获取多有的sheet的名字
    sheet_names = wb.get_sheet_names()
    print(sheet_names)
    # 根据sheet名字获取sheet
    info_sheet = wb.get_sheet_by_name(sheet_names[0])
    # 获取工作簿的最大行数
    # print(info_sheet.max_row)
    # # 获取工作簿的最大列数
    # print(info_sheet.max_column)
    res = []
    for i in range(2, info_sheet.max_row + 1):
        info = {}
        for j in range(1, info_sheet.max_column + 1):
            value = info_sheet.cell(row=i, column=j).value
            print("第{}行第{}列的数据是：{}".format(i, j, value))
            # info_sheet.cell(row=1,column=j).value: key
            info[info_sheet.cell(row=1, column=j).value] = value
        res.append(info)
    return res


info = read_excel("pyth_demo.xlsx")
print(info)
[{"name": "john", "age": 21}, {}, {}, {}]
